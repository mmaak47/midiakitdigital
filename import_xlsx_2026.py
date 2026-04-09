"""
Import VENDAS 2026.xlsx data into the Gestão Comercial system.
Parses monthly sheets (JAN-APR 2026), ACUMULADO META 2026, and RENOVAÇÕES 2026.
"""
import json
import sys
import openpyxl
import requests
from datetime import datetime

# ─── Config ───────────────────────────────────────────────────────────────
XLSX_PATH = r'C:\Users\Usuário\Downloads\VENDAS 2026.xlsx'
API_BASE = 'http://187.127.8.196/api'
ADMIN_USER = 'admin'
ADMIN_PASS = 'intermidia2025'

# Mapping XLSX vendedor short names → system usernames
VENDEDOR_MAP = {
    'EDUARDA': 'eduarda.rossi',
    'JULIANA': 'juliana.kosuta',
    'ESCRITORIO': 'emerson.silva',
    'EMERSON': 'emerson.silva',
}

MONTH_SHEETS_2026 = {
    'JANEIRO 2026': 1,
    'FEVEREIRO 2026': 2,
    'MARÇO 2026': 3,
    'ABRIL 2026': 4,
}

def login():
    r = requests.post(f'{API_BASE}/auth/login', json={'username': ADMIN_USER, 'password': ADMIN_PASS})
    r.raise_for_status()
    return r.json()['token']

def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s or s in ('', '0', '#REF!'):
        return None
    return s

def parse_bool(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ('true', '1', 'yes')

def parse_num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(',', '.')
    if not s or s in ('#REF!', '-'):
        return 0
    try:
        return float(s)
    except ValueError:
        return 0

def parse_vendedor_section(ws, start_row, ano, mes):
    """Parse a vendedor section from a monthly sheet.
    Returns (vendedor_username, meta_base, vendas_list, end_row)"""
    # Row structure:
    # start_row: VENDEDOR_NAME ... META MÊS/HIPER META ... values
    # start_row+1: ... META BASE/RECORRENCIA ... values
    # start_row+2: S T A T U S headers
    # start_row+3: Contrato, Contrato assinado, ...
    # start_row+4...: data rows (until TOTAL or empty section)

    vendedor_raw = str(ws.cell(start_row, 1).value or '').strip().upper()
    if not vendedor_raw:
        return None, 0, [], start_row + 1

    vendedor = VENDEDOR_MAP.get(vendedor_raw, vendedor_raw.lower())

    # Get META BASE (usually in start_row+1, col I=9)
    meta_base = 0
    for offset in [0, 1]:
        label = str(ws.cell(start_row + offset, 8).value or '').strip().upper()
        if 'META' in label and 'BASE' in label:
            meta_base = parse_num(ws.cell(start_row + offset, 9).value)
            break
        if 'META' in label and ('MES' in label or 'MÊS' in label):
            meta_base = parse_num(ws.cell(start_row + offset, 9).value)

    # Data rows start at start_row + 4 (after header + sub-header)
    data_start = start_row + 4
    vendas = []

    for r in range(data_start, data_start + 20):  # max 20 rows per vendor
        c1 = ws.cell(r, 1).value
        # Check for TOTAL row or empty
        if c1 is not None and str(c1).strip().upper().startswith('TOTAL'):
            break
        # Empty row check (col 8=cliente empty means row is empty or padding)
        cliente = str(ws.cell(r, 8).value or '').strip()
        valor = parse_num(ws.cell(r, 11).value)

        # Skip truly empty rows
        if not cliente and valor == 0:
            continue

        if not cliente:
            continue

        venda = {
            'vendedor_nome': vendedor,
            'ano': ano,
            'mes': mes,
            'status_contrato': parse_bool(ws.cell(r, 1).value),
            'status_contrato_assinado': parse_bool(ws.cell(r, 2).value),
            'status_conteudo': parse_bool(ws.cell(r, 3).value),
            'status_checkin': parse_bool(ws.cell(r, 4).value),
            'status_faturado': parse_bool(ws.cell(r, 5).value),
            'status_excel_pastas': parse_bool(ws.cell(r, 6).value),
            'data_venda': parse_date(ws.cell(r, 7).value),
            'cliente': cliente,
            'cnpj': str(ws.cell(r, 9).value or '').strip() or None,
            'pontos_contratados': str(ws.cell(r, 10).value or '').strip() or None,
            'valor_mensal': parse_num(ws.cell(r, 11).value),
            'total_contrato': parse_num(ws.cell(r, 12).value),
            'qtde_parcelas': int(parse_num(ws.cell(r, 13).value)) or 1,
            'previsao_veiculacao': parse_date(ws.cell(r, 14).value),
            'data_emissao_nf': parse_date(ws.cell(r, 15).value),
        }
        vendas.append(venda)

    # Find end row (skip to after TOTAL + blank rows)
    end_row = data_start + 20
    for r in range(data_start, data_start + 25):
        c1 = ws.cell(r, 1).value
        if c1 is not None and str(c1).strip().upper().startswith('TOTAL'):
            end_row = r + 3  # skip TOTAL + 2 blank rows
            break

    return vendedor, meta_base, vendas, end_row

def find_vendedor_sections(ws):
    """Find the starting rows of vendedor sections."""
    sections = []
    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or '').strip().upper()
        if val in VENDEDOR_MAP:
            sections.append(r)
    return sections

def parse_monthly_sheet(wb, sheet_name, mes):
    ws = wb[sheet_name]
    ano = 2026
    all_vendas = []
    all_metas = []

    sections = find_vendedor_sections(ws)
    for sec_row in sections:
        vendedor, meta_base, vendas, _ = parse_vendedor_section(ws, sec_row, ano, mes)
        if vendedor:
            all_vendas.extend(vendas)
            if meta_base > 0:
                all_metas.append({
                    'vendedor_nome': vendedor,
                    'ano': ano,
                    'mes': mes,
                    'valor_meta': meta_base,
                })

    return all_vendas, all_metas

def parse_renovacoes(wb):
    ws = wb['RENOVAÇÕES 2026']
    renovacoes = []
    current_mes = 0

    MONTH_MAP = {
        'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'ABRIL': 4,
        'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8,
        'SETEMBRO': 9, 'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12,
    }

    for r in range(1, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip().upper()

        # Detect month header
        if c1 in MONTH_MAP:
            current_mes = MONTH_MAP[c1]
            continue

        # Skip header rows
        if c1 in ('CLIENTE', 'CLIENTE ', '') or c1.startswith('TOTAL') or c1.startswith('RENOVADO') or c1.startswith('SEM RENOVAÇÃO'):
            continue

        if current_mes == 0:
            continue

        cliente = str(ws.cell(r, 1).value or '').strip()
        valor = parse_num(ws.cell(r, 2).value)
        if not cliente or valor == 0:
            continue

        parcela = str(ws.cell(r, 3).value or '').strip()
        renovado = parse_bool(ws.cell(r, 6).value)
        sem_renovacao = parse_bool(ws.cell(r, 7).value)
        vendedor_raw = str(ws.cell(r, 8).value or '').strip().upper()
        vendedor = VENDEDOR_MAP.get(vendedor_raw, vendedor_raw.lower() if vendedor_raw else None)

        status = 'pendente'
        if renovado:
            status = 'concluida'
        elif sem_renovacao:
            status = 'perdida'

        renovacoes.append({
            'ano': 2026,
            'mes': current_mes,
            'cliente': cliente,
            'valor_mensal': valor,
            'status': status,
            'vendedor_nome': vendedor,
            'obs': f'Parcela: {parcela}' if parcela else None,
        })

    return renovacoes

def main():
    print(f'Abrindo {XLSX_PATH}...')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_vendas = []
    all_metas = []

    # Parse monthly sheets
    for sheet_name, mes in MONTH_SHEETS_2026.items():
        if sheet_name not in wb.sheetnames:
            print(f'  [SKIP] Sheet "{sheet_name}" não encontrada')
            continue
        vendas, metas = parse_monthly_sheet(wb, sheet_name, mes)
        all_vendas.extend(vendas)
        all_metas.extend(metas)
        print(f'  {sheet_name}: {len(vendas)} vendas, {len(metas)} metas')

    # Parse renovações
    renovacoes = parse_renovacoes(wb)
    print(f'  RENOVAÇÕES 2026: {len(renovacoes)} registros')

    # Summary
    print(f'\nTotal: {len(all_vendas)} vendas, {len(all_metas)} metas, {len(renovacoes)} renovações')

    if not all_vendas and not all_metas and not renovacoes:
        print('Nenhum dado para importar.')
        return

    # Login
    print('\nAutenticando...')
    token = login()
    print('  OK')

    # Send to API
    print('Enviando dados...')
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
    }
    payload = {
        'vendas': all_vendas,
        'metas': all_metas,
        'renovacoes': renovacoes,
    }

    r = requests.post(f'{API_BASE}/gestao/import', headers=headers, json=payload, timeout=60)
    r.raise_for_status()
    result = r.json()
    print(f'  Importado: {result.get("importedVendas", 0)} vendas, '
          f'{result.get("importedMetas", 0)} metas, '
          f'{result.get("importedRenovacoes", 0)} renovações')
    print('\nImportação concluída!')

if __name__ == '__main__':
    main()
